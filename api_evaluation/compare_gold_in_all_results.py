"""
Compare: was the gold-standard website/email found anywhere in all_urls / all_emails
(rather than only in the top result)?
"""
import json
import csv
from urllib.parse import urlparse

try:
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill
    from openpyxl.utils import get_column_letter
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False


def normalize_url(u):
    if not u or not isinstance(u, str):
        return ""
    u = u.strip().lower()
    u = u.replace("https://", "").replace("http://", "")
    u = u.replace("www.", "")
    u = u.rstrip("/")
    return u


def gold_website_in_urls(gold_website, all_urls):
    """True if gold website matches or is contained in any of all_urls (or vice versa)."""
    if not gold_website or not all_urls:
        return False
    g = normalize_url(gold_website)
    if not g:
        return False
    for url in all_urls:
        if not url:
            continue
        n = normalize_url(url)
        if g == n or g in n or n in g:
            return True
    return False


def gold_email_in_emails(gold_email, all_emails):
    """True if gold email appears in all_emails (case-insensitive)."""
    if not gold_email or not all_emails:
        return False
    g = gold_email.strip().lower()
    for e in all_emails:
        if e and e.strip().lower() == g:
            return True
    return False


def load_results(path):
    with open(path) as f:
        return json.load(f)["results"]


def main():
    exa = load_results("results_affiliation_only/exa_results.json")
    tavily = load_results("results_affiliation_only/tavily_results.json")
    assert len(exa) == len(tavily), "Mismatch in number of results"

    rows = []
    exa_web_yes = exa_email_yes = 0
    tavily_web_yes = tavily_email_yes = 0
    n_with_gold_web = 0
    n_with_gold_email = 0

    for i, (e, t) in enumerate(zip(exa, tavily)):
        name = e["name"]
        affiliation = e["affiliation"]
        gold_web = e.get("gold_website", "")
        gold_email = e.get("gold_email", "")

        if gold_web:
            n_with_gold_web += 1
        if gold_email:
            n_with_gold_email += 1

        exa_web_in = gold_website_in_urls(gold_web, e.get("all_urls", []))
        exa_email_in = gold_email_in_emails(gold_email, e.get("all_emails", []))
        tavily_web_in = gold_website_in_urls(gold_web, t.get("all_urls", []))
        tavily_email_in = gold_email_in_emails(gold_email, t.get("all_emails", []))

        if exa_web_in:
            exa_web_yes += 1
        if exa_email_in:
            exa_email_yes += 1
        if tavily_web_in:
            tavily_web_yes += 1
        if tavily_email_in:
            tavily_email_yes += 1

        # Gold website MISSING = has gold URL but not in API's all_urls (highlight these)
        exa_web_missing = bool(gold_web and not exa_web_in)
        tavily_web_missing = bool(gold_web and not tavily_web_in)
        rows.append({
            "name": name,
            "affiliation": affiliation,
            "gold_website": gold_web,
            "gold_email": gold_email,
            "exa_gold_website_in_all_urls": exa_web_in,
            "exa_gold_website_MISSING": exa_web_missing,
            "exa_gold_email_in_all_emails": exa_email_in,
            "tavily_gold_website_in_all_urls": tavily_web_in,
            "tavily_gold_website_MISSING": tavily_web_missing,
            "tavily_gold_email_in_all_emails": tavily_email_in,
            "_sort_missing_first": exa_web_missing or tavily_web_missing,
        })

    # Sort: rows where gold website is NOT in all_urls come first (so they're easy to see)
    rows.sort(key=lambda r: (not r["_sort_missing_first"], r["name"]))
    for r in rows:
        del r["_sort_missing_first"]

    # Write CSV (missing columns make it easy to filter/highlight in Excel)
    out_path = "results_affiliation_only/gold_in_all_results_comparison.csv"
    fieldnames = [
        "name", "affiliation", "gold_website", "gold_email",
        "exa_gold_website_in_all_urls", "exa_gold_website_MISSING",
        "exa_gold_email_in_all_emails",
        "tavily_gold_website_in_all_urls", "tavily_gold_website_MISSING",
        "tavily_gold_email_in_all_emails",
    ]
    with open(out_path, "w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=fieldnames)
        w.writeheader()
        w.writerows(rows)
    print("Wrote:", out_path)

    # Excel version with highlighted rows (red = gold website NOT in all_urls)
    if HAS_OPENPYXL:
        xlsx_path = "results_affiliation_only/gold_in_all_results_comparison.xlsx"
        wb = Workbook()
        ws = wb.active
        ws.title = "Gold in all_urls comparison"
        highlight_fill = PatternFill(start_color="FFCCCB", end_color="FFCCCB", fill_type="solid")  # light red
        for col, key in enumerate(fieldnames, 1):
            ws.cell(row=1, column=col, value=key)
        for row_idx, r in enumerate(rows, 2):
            for col_idx, key in enumerate(fieldnames, 1):
                val = r.get(key, "")
                if isinstance(val, bool):
                    val = "Yes" if val else "No"
                ws.cell(row=row_idx, column=col_idx, value=val)
            if r.get("exa_gold_website_MISSING") or r.get("tavily_gold_website_MISSING"):
                for col in range(1, len(fieldnames) + 1):
                    ws.cell(row=row_idx, column=col).fill = highlight_fill
        wb.save(xlsx_path)
        print("Wrote (highlighted):", xlsx_path)

    # Summary report
    n = len(rows)
    report_path = "results_affiliation_only/gold_in_all_results_report.md"
    with open(report_path, "w") as f:
        f.write("# Gold standard: found anywhere in all_urls / all_emails?\n\n")
        f.write("Comparison of whether the **gold-standard** website or email appears **anywhere** in the full result set (all_urls, all_emails), not just in the top hit.\n\n")
        f.write("**Professors with gold website:** " + str(n_with_gold_web) + "\n\n")
        f.write("**Professors with gold email:** " + str(n_with_gold_email) + "\n\n")
        f.write("## Summary\n\n")
        f.write("| Metric | Exa | Tavily |\n")
        f.write("|--------|-----|--------|\n")
        f.write(f"| Gold website found in all_urls (of {n_with_gold_web} with gold) | {exa_web_yes} ({100*exa_web_yes/max(n_with_gold_web,1):.1f}%) | {tavily_web_yes} ({100*tavily_web_yes/max(n_with_gold_web,1):.1f}%) |\n")
        f.write(f"| Gold email found in all_emails (of {n_with_gold_email} with gold) | {exa_email_yes} ({100*exa_email_yes/max(n_with_gold_email,1):.1f}%) | {tavily_email_yes} ({100*tavily_email_yes/max(n_with_gold_email,1):.1f}%) |\n")
        f.write("\n## Interpretation\n\n")
        f.write("- **Gold website in all_urls:** The gold-standard URL (or a normalized match) appeared in at least one of the URLs returned by the API.\n")
        f.write("- **Gold email in all_emails:** The gold-standard email appeared in the list of emails extracted from the API results.\n")
    print("Wrote:", report_path)

    print("\n--- Summary ---")
    print(f"Professors with gold website: {n_with_gold_web}")
    print(f"Professors with gold email:  {n_with_gold_email}")
    print(f"Exa   – gold website in all_urls:  {exa_web_yes}/{n_with_gold_web} ({100*exa_web_yes/max(n_with_gold_web,1):.1f}%)")
    print(f"Exa   – gold email in all_emails:  {exa_email_yes}/{n_with_gold_email} ({100*exa_email_yes/max(n_with_gold_email,1):.1f}%)")
    print(f"Tavily – gold website in all_urls:  {tavily_web_yes}/{n_with_gold_web} ({100*tavily_web_yes/max(n_with_gold_web,1):.1f}%)")
    print(f"Tavily – gold email in all_emails:  {tavily_email_yes}/{n_with_gold_email} ({100*tavily_email_yes/max(n_with_gold_email,1):.1f}%)")


if __name__ == "__main__":
    main()
